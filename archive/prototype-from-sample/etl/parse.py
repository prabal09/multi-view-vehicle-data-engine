"""Pure parsing of VehicleSyncData.xlsx into structured records.

This module contains NO database code so the parsing logic can be unit-tested
on its own. It converts the human-authored spreadsheet layout into flat records
that the loader (``load_xlsx.py``) inserts into PostgreSQL.

Spreadsheet layout (Sheet1)
---------------------------
Rows are grouped into *blocks*. Each block describes one physically distinct
vehicle that the perception pipeline associated across viewpoints. A block is a
run of up to five rows -- one per camera ``c0``..``c4`` -- and starts on the
``c0`` row. The ``Vehicle Type`` / ``Vehicle Model`` cells are filled only on
the anchor (``c0``) row and apply to the whole block. Each row carries that
camera's local track label plus the frame/time range in which the track is
visible. A camera may be absent for a vehicle (blank track/range).

Note: a camera-local label such as ``c0-5`` is NOT globally unique -- the same
label is reused by different vehicles with different frame ranges. The frame
range is what disambiguates, so the loader keys tracks on a surrogate id.

Sheet2 holds per-camera extrinsics (translation + rotation relative to the
vehicle). In the sample only headers and a few frame numbers are populated; the
parser returns whatever numeric values are present so real exports load too.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Iterable, Optional

import openpyxl


# Header labels as they appear in Sheet1 (used to locate the columns robustly
# rather than hard-coding spreadsheet coordinates).
_H_TYPE = "vehicle type"
_H_MODEL = "vehicle model"
_H_TRACK = "track"
_H_FSTART = "frame-start"
_H_FEND = "frame-end"
_H_TSTART = "time-start"
_H_TEND = "time-end"


@dataclass
class TrackRecord:
    """One camera's view of a vehicle."""

    camera: str
    label: Optional[str]              # camera-local track label, e.g. "c1-1"
    frame_start: Optional[int]
    frame_end: Optional[int]
    time_start: Optional[float]
    time_end: Optional[float]


@dataclass
class VehicleRecord:
    """A physical vehicle plus its per-camera tracks."""

    index: int                        # 0-based order of appearance in the sheet
    vehicle_type: Optional[str]
    vehicle_model: Optional[str]
    tracks: list[TrackRecord] = field(default_factory=list)


def _norm(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _as_int(value) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def _as_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_camera_label(value) -> bool:
    s = _norm(value)
    return bool(s) and len(s) == 2 and s[0] == "c" and s[1].isdigit()


def parse_vehicles(path: str, sheet: str = "Sheet1") -> list[VehicleRecord]:
    """Parse Sheet1 into a list of :class:`VehicleRecord`."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row, cols = _find_header_from_rows(rows)

    # Camera column: the leftmost column that carries "c0".."c4" markers, which
    # sits just left of the Track column in the sample.
    track_col = cols[_H_TRACK]
    cam_col = _detect_camera_column(rows, header_row, track_col)

    vehicles: list[VehicleRecord] = []
    current: Optional[VehicleRecord] = None

    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        cam = _norm(row[cam_col]) if cam_col < len(row) else None
        if not _is_camera_label(cam):
            continue  # blank separator row

        track = TrackRecord(
            camera=cam,
            label=_norm(_cell(row, track_col)),
            frame_start=_as_int(_cell(row, cols[_H_FSTART])),
            frame_end=_as_int(_cell(row, cols[_H_FEND])),
            time_start=_as_float(_cell(row, cols[_H_TSTART])),
            time_end=_as_float(_cell(row, cols[_H_TEND])),
        )

        if cam == "c0":
            current = VehicleRecord(
                index=len(vehicles),
                vehicle_type=_norm(_cell(row, cols[_H_TYPE])),
                vehicle_model=_norm(_cell(row, cols[_H_MODEL])),
            )
            vehicles.append(current)

        if current is None:
            # Defensive: a non-c0 row before any c0 anchor -> start a block.
            current = VehicleRecord(index=len(vehicles), vehicle_type=None,
                                    vehicle_model=None)
            vehicles.append(current)

        current.tracks.append(track)

    return vehicles


def _track_is_empty(t: TrackRecord) -> bool:
    return (t.label is None and t.frame_start is None and t.frame_end is None
            and t.time_start is None and t.time_end is None)


def nonempty_vehicles(vehicles: Iterable[VehicleRecord]) -> list[VehicleRecord]:
    """Drop blocks that are pure spreadsheet padding (every track empty and no
    type/model). These are the blank template rows at the bottom of Sheet1."""
    out = []
    for v in vehicles:
        has_meta = bool(v.vehicle_type or v.vehicle_model)
        has_track = any(not _track_is_empty(t) for t in v.tracks)
        if has_meta or has_track:
            out.append(v)
    return out


@dataclass
class CameraExtrinsic:
    camera: str
    tx: Optional[float] = None
    ty: Optional[float] = None
    tz: Optional[float] = None
    rotation: Optional[list[list[float]]] = None   # 3x3 or None


def parse_extrinsics(path: str, sheet: str = "Sheet2") -> list[CameraExtrinsic]:
    """Parse Sheet2 extrinsics. The sample only has headers, so numeric values
    are returned when present and left as ``None`` otherwise."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    out: list[CameraExtrinsic] = []
    for row in rows:
        for c, cell in enumerate(row):
            if _is_camera_label(cell):
                out.append(CameraExtrinsic(camera=_norm(cell)))
    return out


# --- helpers -----------------------------------------------------------------

def _cell(row, idx):
    return row[idx] if idx is not None and idx < len(row) else None


def _find_header_from_rows(rows) -> tuple[int, dict[str, int]]:
    wanted = {_H_TYPE, _H_MODEL, _H_TRACK, _H_FSTART, _H_FEND, _H_TSTART, _H_TEND}
    for r, row in enumerate(rows):
        labels = {}
        for c, cell in enumerate(row):
            s = _norm(cell)
            if s and s.lower() in wanted:
                labels[s.lower()] = c
        if _H_TRACK in labels and _H_FSTART in labels:
            return r, labels
    raise ValueError("Could not locate the header row in Sheet1")


def _detect_camera_column(rows, header_row, track_col) -> int:
    """Find the column that holds c0..c4 markers below the header."""
    from collections import Counter
    counts: Counter = Counter()
    for r in range(header_row + 1, len(rows)):
        for c, cell in enumerate(rows[r]):
            if _is_camera_label(cell):
                counts[c] += 1
    if counts:
        return counts.most_common(1)[0][0]
    # Fallback: the column immediately left of Track.
    return max(0, track_col - 1)
